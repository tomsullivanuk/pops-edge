import math
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    BET_LOG_FILE,
    COL_ACTION,
    COL_CONTRACTS,
    COL_ENTRY_PRICE,
    COL_MARKET_TICKER,
    COL_MARKET_TICKER_TITLE,
    COL_STATUS,
    FUTURES_VALUE_BOARD_FILE,
    LADDER_BOARD_FILE,
    SHEET_DEFINITIONS,
    SHEET_RUN_METADATA,
    SHEET_VALUE_BOARD,
    SHEET_WAGERS,
    VALUE_BOARD_FILE,
)

ACTIVE_STATUSES = ["Open", "Partially Closed"]
SHEET_LADDER_BOARD = "Ladder Board"
EXIT_CONTRACTS_COLUMN = "Exit Contracts"
REMAINING_CONTRACTS_COLUMN = "Remaining Contracts"

OUTPUT_COLUMNS = [
    "event_title",
    "market_title",
    "side",
    "contracts_held",
    "avg_price",
    "current_price",
    "silver_probability",
    "current_edge",
    "ladder_1_price",
    "ladder_1_contracts",
    "ladder_2_price",
    "ladder_2_contracts",
    "ladder_3_price",
    "ladder_3_contracts",
    "ladder_display",
    "estimated_gross_proceeds",
    "estimated_profit_if_all_filled",
    "market_ticker",
    "market_type",
    "status",
    "action_flag",
]

VALUE_ROW_COLUMNS = [
    COL_MARKET_TICKER,
    "event_title",
    "market_title",
    "yes_bid",
    "yes_ask",
    "silver_probability_yes",
    "market_type",
]


def make_unique_columns(columns):
    counts = {}
    unique_columns = []
    for column in columns:
        column_name = str(column)
        counts[column_name] = counts.get(column_name, 0) + 1
        if counts[column_name] == 1:
            unique_columns.append(column_name)
        else:
            unique_columns.append(f"{column_name}__{counts[column_name]}")
    return unique_columns


def numeric(value):
    return pd.to_numeric(value, errors="coerce")


def round_cent(value):
    if pd.isna(value):
        return pd.NA
    rounded = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(min(rounded, Decimal("0.99")))


def format_cents(value):
    if pd.isna(value) or value == "":
        return ""
    return f"{int(round(float(value) * 100))}¢"


def allocate_contracts(contracts_held):
    contracts = int(contracts_held)
    tier_1 = math.floor(contracts * 0.30)
    tier_2 = math.floor(contracts * 0.30)
    tier_3 = contracts - tier_1 - tier_2
    return tier_1, tier_2, tier_3


def calculate_ladder(entry_price, fair_value, contracts_held):
    entry = numeric(entry_price)
    fair = numeric(fair_value)
    contracts = numeric(contracts_held)

    empty = {
        "current_edge": pd.NA,
        "ladder_1_price": pd.NA,
        "ladder_1_contracts": pd.NA,
        "ladder_2_price": pd.NA,
        "ladder_2_contracts": pd.NA,
        "ladder_3_price": pd.NA,
        "ladder_3_contracts": pd.NA,
        "ladder_display": "",
        "estimated_gross_proceeds": pd.NA,
        "estimated_profit_if_all_filled": pd.NA,
    }

    if pd.isna(entry) or pd.isna(fair) or pd.isna(contracts) or contracts <= 0:
        return empty

    edge = fair - entry
    if edge <= 0:
        return {**empty, "current_edge": round(edge, 4)}

    whole_contracts = int(contracts)
    ladder_prices = [
        round_cent(entry + 0.50 * edge),
        round_cent(fair),
        round_cent(fair + 0.15),
    ]
    ladder_contracts = allocate_contracts(whole_contracts)
    gross_proceeds = round(
        sum(price * count for price, count in zip(ladder_prices, ladder_contracts)),
        2,
    )
    profit = round(gross_proceeds - (whole_contracts * entry), 2)
    display = " | ".join(
        f"{format_cents(price)} × {count}"
        for price, count in zip(ladder_prices, ladder_contracts)
    )

    return {
        "current_edge": round(edge, 4),
        "ladder_1_price": ladder_prices[0],
        "ladder_1_contracts": ladder_contracts[0],
        "ladder_2_price": ladder_prices[1],
        "ladder_2_contracts": ladder_contracts[1],
        "ladder_3_price": ladder_prices[2],
        "ladder_3_contracts": ladder_contracts[2],
        "ladder_display": display,
        "estimated_gross_proceeds": gross_proceeds,
        "estimated_profit_if_all_filled": profit,
    }


def read_active_wagers(path=BET_LOG_FILE):
    try:
        wagers = pd.read_excel(path, sheet_name=SHEET_WAGERS)
    except FileNotFoundError:
        return pd.DataFrame()

    required = [
        COL_MARKET_TICKER_TITLE,
        COL_ACTION,
        COL_ENTRY_PRICE,
        COL_CONTRACTS,
    ]
    missing = [column for column in required if column not in wagers.columns]
    if missing:
        print(f"WARNING: Could not build ladder board; missing wager columns: {missing}")
        return pd.DataFrame()

    wagers = wagers.copy()
    if COL_STATUS not in wagers.columns:
        wagers[COL_STATUS] = ""

    wagers = wagers[wagers[COL_STATUS].isin(ACTIVE_STATUSES)].copy()
    if wagers.empty:
        return wagers

    contracts = pd.to_numeric(wagers[COL_CONTRACTS], errors="coerce")
    if EXIT_CONTRACTS_COLUMN in wagers.columns:
        exit_contracts = pd.to_numeric(
            wagers[EXIT_CONTRACTS_COLUMN],
            errors="coerce",
        ).fillna(0)
        calculated_remaining = (contracts - exit_contracts).clip(lower=0)
    else:
        calculated_remaining = contracts

    if REMAINING_CONTRACTS_COLUMN in wagers.columns:
        recorded_remaining = pd.to_numeric(
            wagers[REMAINING_CONTRACTS_COLUMN],
            errors="coerce",
        )
        wagers[REMAINING_CONTRACTS_COLUMN] = recorded_remaining.fillna(
            calculated_remaining,
        )
    else:
        wagers[REMAINING_CONTRACTS_COLUMN] = calculated_remaining

    wagers = wagers[pd.to_numeric(wagers[REMAINING_CONTRACTS_COLUMN], errors="coerce").fillna(0) > 0].copy()
    if wagers.empty:
        return wagers

    wagers["_side"] = wagers[COL_ACTION].map({"BUY YES": "YES", "SELL YES": "NO"}).fillna(wagers[COL_ACTION])
    wagers["_weighted_entry"] = (
        pd.to_numeric(wagers[COL_ENTRY_PRICE], errors="coerce").fillna(0)
        * pd.to_numeric(wagers[REMAINING_CONTRACTS_COLUMN], errors="coerce").fillna(0)
    )

    grouped_rows = []
    for (market_ticker, action, side), group in wagers.groupby(
        [COL_MARKET_TICKER_TITLE, COL_ACTION, "_side"],
        dropna=False,
    ):
        remaining_contracts = pd.to_numeric(
            group[REMAINING_CONTRACTS_COLUMN],
            errors="coerce",
        ).fillna(0).sum()
        if remaining_contracts <= 0:
            continue

        weighted_entry = group["_weighted_entry"].sum()
        avg_entry = weighted_entry / remaining_contracts if remaining_contracts else pd.NA
        statuses = [
            str(status)
            for status in group[COL_STATUS].dropna().unique()
            if str(status).strip()
        ]
        grouped_rows.append({
            COL_MARKET_TICKER_TITLE: market_ticker,
            COL_ACTION: action,
            COL_ENTRY_PRICE: avg_entry,
            COL_CONTRACTS: remaining_contracts,
            COL_STATUS: ", ".join(statuses),
        })

    return pd.DataFrame(grouped_rows)


def read_value_board_rows(path, market_type):
    try:
        frame = pd.read_excel(path, sheet_name=SHEET_VALUE_BOARD)
    except FileNotFoundError:
        return pd.DataFrame(columns=VALUE_ROW_COLUMNS)

    return normalize_value_board_rows(frame, market_type)


def first_column(frame, candidates):
    for candidate in candidates:
        if candidate in frame.columns:
            return frame[candidate]
    return pd.Series(pd.NA, index=frame.index)


def normalize_value_board_rows(frame, market_type):
    frame = frame.copy()
    frame.columns = make_unique_columns(frame.columns)
    assert frame.columns.is_unique

    canonical = pd.DataFrame(index=frame.index)
    canonical[COL_MARKET_TICKER] = first_column(frame, [COL_MARKET_TICKER, COL_MARKET_TICKER_TITLE])
    canonical["event_title"] = first_column(frame, ["game", "event_title", "Stage", "Event"])
    canonical["market_title"] = first_column(frame, ["market_title", "Market Title"])
    canonical["yes_bid"] = first_column(frame, ["yes_bid", "Yes Bid"])
    canonical["yes_ask"] = first_column(frame, ["yes_ask", "Yes Ask"])
    canonical["silver_probability_yes"] = first_column(
        frame,
        ["silver_prob", "silver_probability_yes", "Silver", "Silver Probability"],
    )
    canonical["market_type"] = market_type
    canonical = canonical[VALUE_ROW_COLUMNS].copy()
    assert canonical.columns.is_unique
    return canonical


def build_ladder_board(active_wagers, value_rows):
    if active_wagers.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    lookup = value_rows.rename(columns={COL_MARKET_TICKER: COL_MARKET_TICKER_TITLE})
    positions = active_wagers.merge(lookup, on=COL_MARKET_TICKER_TITLE, how="left")

    rows = []
    for _, row in positions.iterrows():
        action = row[COL_ACTION]
        entry_price_yes = numeric(row[COL_ENTRY_PRICE])
        contracts = numeric(row[COL_CONTRACTS])
        yes_bid = numeric(row["yes_bid"])
        yes_ask = numeric(row["yes_ask"])
        silver_yes = numeric(row["silver_probability_yes"])

        if action == "BUY YES":
            side = "YES"
            avg_price = entry_price_yes
            current_price = yes_bid
            fair_value = silver_yes
        else:
            side = "NO"
            avg_price = 1 - entry_price_yes if pd.notna(entry_price_yes) else pd.NA
            current_price = 1 - yes_ask if pd.notna(yes_ask) else pd.NA
            fair_value = 1 - silver_yes if pd.notna(silver_yes) else pd.NA

        ladder = calculate_ladder(avg_price, fair_value, contracts)
        action_flag = ""
        if pd.notna(current_price) and pd.notna(ladder["ladder_1_price"]):
            if current_price >= ladder["ladder_1_price"]:
                action_flag = "urgent"
            elif round(ladder["ladder_1_price"] - current_price, 2) <= 0.03:
                action_flag = "near"

        rows.append({
            "event_title": row.get("event_title", ""),
            "market_title": row.get("market_title", ""),
            "side": side,
            "contracts_held": int(contracts) if pd.notna(contracts) else pd.NA,
            "avg_price": round(avg_price, 4) if pd.notna(avg_price) else pd.NA,
            "current_price": round(current_price, 4) if pd.notna(current_price) else pd.NA,
            "silver_probability": round(fair_value, 4) if pd.notna(fair_value) else pd.NA,
            **ladder,
            "market_ticker": row[COL_MARKET_TICKER_TITLE],
            "market_type": row.get("market_type", ""),
            "status": row[COL_STATUS],
            "action_flag": action_flag,
        })

    ladder_board = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    priority = {"urgent": 0, "near": 1, "": 2}
    ladder_board["_sort_priority"] = ladder_board["action_flag"].map(priority).fillna(2)
    return ladder_board.sort_values(
        by=[
            "_sort_priority",
            "current_edge",
            "event_title",
            "market_title",
            "market_ticker",
            "side",
        ],
        ascending=[True, False, True, True, True, True],
        na_position="last",
    ).drop(columns=["_sort_priority"])


def write_workbook(ladder_board, output_file=LADDER_BOARD_FILE):
    metadata = pd.DataFrame([
        ["Ladder Board Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Open Positions", len(ladder_board)],
        ["Rows With Ladders", ladder_board["ladder_display"].astype(bool).sum() if not ladder_board.empty else 0],
        ["Position Filter", "Remaining Contracts greater than zero"],
    ], columns=["Field", "Value"])
    definitions = pd.DataFrame([
        ["side", "YES positions use YES prices and Silver probability. NO positions use NO prices and one minus Silver probability."],
        ["current_edge", "Side-specific fair value minus side-specific average entry price."],
        ["ladder_1_price", "Entry price plus 50% of positive edge, capped at 99 cents."],
        ["ladder_2_price", "Side-specific Silver fair value, capped at 99 cents."],
        ["ladder_3_price", "Side-specific Silver fair value plus 15 cents, capped at 99 cents."],
        ["action_flag", "urgent when current price is at or above tier 1; near when tier 1 is within 3 cents."],
    ], columns=["Field", "Plain-English Definition"])

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        ladder_board.to_excel(writer, sheet_name=SHEET_LADDER_BOARD, index=False)
        metadata.to_excel(writer, sheet_name=SHEET_RUN_METADATA, index=False)
        definitions.to_excel(writer, sheet_name=SHEET_DEFINITIONS, index=False)

        worksheet = writer.sheets[SHEET_LADDER_BOARD]
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        urgent_fill = PatternFill("solid", fgColor="FEE2E2")
        near_fill = PatternFill("solid", fgColor="FEF3C7")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)

        headers = [cell.value for cell in worksheet[1]]
        action_flag_index = headers.index("action_flag") + 1
        money_columns = {
            "avg_price",
            "current_price",
            "silver_probability",
            "current_edge",
            "ladder_1_price",
            "ladder_2_price",
            "ladder_3_price",
        }
        currency_columns = {"estimated_gross_proceeds", "estimated_profit_if_all_filled"}

        for row in worksheet.iter_rows(min_row=2):
            flag = row[action_flag_index - 1].value
            if flag == "urgent":
                fill = urgent_fill
            elif flag == "near":
                fill = near_fill
            else:
                fill = None
            for cell in row:
                if fill:
                    cell.fill = fill

        for column_index, header in enumerate(headers, start=1):
            letter = get_column_letter(column_index)
            if header in money_columns:
                for cell in worksheet[letter][1:]:
                    cell.number_format = "0%"
            elif header in currency_columns:
                for cell in worksheet[letter][1:]:
                    cell.number_format = "$0.00"
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in worksheet[letter]
            )
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 48)


def main():
    print("Building ladder board...")
    active_wagers = read_active_wagers()
    match_value_rows = read_value_board_rows(VALUE_BOARD_FILE, "Match")
    futures_value_rows = read_value_board_rows(FUTURES_VALUE_BOARD_FILE, "Futures")
    assert match_value_rows.columns.is_unique
    assert futures_value_rows.columns.is_unique
    value_rows = pd.concat([match_value_rows, futures_value_rows], ignore_index=True)
    ladder_board = build_ladder_board(active_wagers, value_rows)
    write_workbook(ladder_board)
    print(f"Saved: {LADDER_BOARD_FILE}")
    print(f"Open positions: {len(ladder_board)}")
    print(f"Rows with ladders: {ladder_board['ladder_display'].astype(bool).sum() if not ladder_board.empty else 0}")


if __name__ == "__main__":
    main()
