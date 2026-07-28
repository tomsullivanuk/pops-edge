import os
import runpy
import shutil
import sys
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures"


class PipelineTests(unittest.TestCase):
    def test_process_silver_builds_clean_workbook_from_sample_csv(self):
        with TemporaryDirectory() as home:
            project_dir = Path(home) / "kalshi"
            downloads_dir = Path(home) / "Downloads"
            project_dir.mkdir()
            downloads_dir.mkdir()
            match_csv = downloads_dir / "data-match.csv"
            futures_csv = downloads_dir / "data-futures.csv"
            shutil.copy2(FIXTURES / "sample_silver.csv", match_csv)
            shutil.copy2(FIXTURES / "sample_silver_futures.csv", futures_csv)
            os.utime(match_csv, (1_700_000_000, 1_700_000_000))
            os.utime(futures_csv, (1_700_000_100, 1_700_000_100))

            with patched_environ(HOME=home):
                run_pipeline_script(ROOT / "process_silver.py")

            output = project_dir / "Silver_Current.xlsx"
            self.assertTrue(output.exists())

            forecasts = pd.read_excel(output, sheet_name="Silver Forecasts")
            metadata = pd.read_excel(output, sheet_name="Silver Metadata")

            self.assertEqual(len(forecasts), 2)
            self.assertEqual(forecasts["Team"].tolist(), ["USA", "CAN"])
            self.assertEqual(forecasts["Opponent"].tolist(), ["MEX", "BRA"])
            self.assertEqual(metadata.loc[metadata["Field"] == "Silver Rows Output", "Value"].iloc[0], 2)
            self.assertEqual(
                metadata.loc[metadata["Field"] == "Silver Source File", "Value"].iloc[0],
                "data-match.csv",
            )
            self.assertTrue(futures_csv.exists())
            self.assertTrue((project_dir / "archive" / "silver_processed").exists())

    def test_process_silver_futures_builds_clean_workbook_from_sample_csv(self):
        with TemporaryDirectory() as home:
            project_dir = Path(home) / "kalshi"
            downloads_dir = Path(home) / "Downloads"
            project_dir.mkdir()
            downloads_dir.mkdir()
            futures_csv = downloads_dir / "data-futures.csv"
            match_csv = downloads_dir / "data-match.csv"
            futures_fixture = pd.read_csv(FIXTURES / "sample_silver_futures.csv")
            futures_fixture = futures_fixture.rename(
                columns={"Champ 🏆": "Champ ðŸ†"},
            )
            futures_fixture.to_csv(futures_csv, index=False)
            shutil.copy2(FIXTURES / "sample_silver.csv", match_csv)
            os.utime(futures_csv, (1_700_000_000, 1_700_000_000))
            os.utime(match_csv, (1_700_000_100, 1_700_000_100))

            with patched_environ(HOME=home):
                run_pipeline_script(ROOT / "process_silver_futures.py")

            output = project_dir / "Silver_Futures_Current.xlsx"
            self.assertTrue(output.exists())

            futures = pd.read_excel(output, sheet_name="Silver Futures")
            metadata = pd.read_excel(output, sheet_name="Silver Metadata")

            self.assertEqual(len(futures), 4)
            self.assertEqual(futures["Team"].tolist(), ["ARG", "USA", "BRA", "SCO"])
            self.assertAlmostEqual(futures.loc[futures["Team"] == "ARG", "Champ"].iloc[0], 0.24113)
            self.assertAlmostEqual(futures.loc[futures["Team"] == "ARG", "R16"].iloc[0], 0.92284)
            self.assertAlmostEqual(futures.loc[futures["Team"] == "USA", "R16"].iloc[0], 0.85356)
            scotland = futures.loc[futures["Team"] == "SCO"].iloc[0]
            expected_scotland = {
                "R32": 0.86341,
                "R16": 0.14066,
                "Qtr": 0.03408,
                "Semi": 0.00813,
                "Final": 0.002,
                "Champ": 0.00047,
                "3rd": 0.0014,
            }
            for column, expected in expected_scotland.items():
                self.assertAlmostEqual(scotland[column], expected)
            self.assertEqual(
                metadata.loc[metadata["Field"] == "Silver Futures Rows Output", "Value"].iloc[0],
                4,
            )
            self.assertEqual(
                metadata.loc[metadata["Field"] == "Silver Futures Source File", "Value"].iloc[0],
                "data-futures.csv",
            )
            self.assertEqual(
                metadata.loc[metadata["Field"] == "Champion Column Used", "Value"].iloc[0],
                "Champ ðŸ†",
            )
            self.assertTrue(match_csv.exists())
            self.assertEqual(len(list((project_dir / "archive" / "silver_futures_raw").glob("*.csv"))), 1)
            self.assertTrue((project_dir / "archive" / "silver_futures_processed").exists())

    def test_kalshi_pull_includes_match_and_world_cup_futures_markets(self):
        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)

            with changed_dir(workspace), patch("requests.get") as get:
                get.return_value = fake_kalshi_response({
                    "events": [
                        kalshi_event(
                            "KXWCGAME-26JUN13-USAMEX",
                            "KXWCGAME",
                            "USA vs Mexico",
                            [
                                kalshi_market("KXWCGAME-26JUN13-USAMEX-USA", "United States", "USA vs Mexico Winner?"),
                            ],
                        ),
                        kalshi_event(
                            "KXWCROUND-26RO16",
                            "KXWCROUND",
                            "World Soccer Cup Round of 16 Qualifiers",
                            [
                                kalshi_market("KXWCROUND-26RO16-USA", "USA", "Will USA qualify for FIFA World Cup Round of 16?"),
                            ],
                        ),
                        kalshi_event(
                            "KXWCROUND-26QUAR",
                            "KXWCROUND",
                            "World Soccer Cup Quarterfinals Qualifiers",
                            [
                                kalshi_market("KXWCROUND-26QUAR-USA", "USA", "Will USA qualify for FIFA World Cup Quarterfinals?"),
                            ],
                        ),
                        kalshi_event(
                            "KXWCROUND-26SEMI",
                            "KXWCROUND",
                            "World Soccer Cup Semifinals Qualifiers",
                            [
                                kalshi_market("KXWCROUND-26SEMI-USA", "USA", "Will USA qualify for FIFA World Cup Semifinals?"),
                            ],
                        ),
                        kalshi_event(
                            "KXMENWORLDCUP-26",
                            "KXMENWORLDCUP",
                            "2026 World Soccer Cup Winner",
                            [
                                kalshi_market("KXMENWORLDCUP-26-AR", "Argentina", "Will the Argentina win the 2026 Men's World Cup?"),
                            ],
                        ),
                        kalshi_event(
                            "KXWCROUND-26FINAL",
                            "KXWCROUND",
                            "World Soccer Cup Final Qualifiers",
                            [
                                kalshi_market("KXWCROUND-26FINAL-USA", "USA", "Will USA qualify for FIFA World Cup Final?"),
                            ],
                        ),
                        kalshi_event(
                            "KXWCHOST-2038",
                            "KXWCHOST",
                            "Who will host the 2038 World Soccer Cup?",
                            [
                                kalshi_market("KXWCHOST-2038-GER", "Germany", "Will Germany host?"),
                            ],
                        ),
                    ],
                    "cursor": None,
                })

                run_pipeline_script(ROOT / "kalshi_pull.py")

            output = workspace / "Kalshi_Current.xlsx"
            self.assertTrue(output.exists())

            pulled = pd.read_excel(output)
            self.assertEqual(
                pulled["market_ticker"].tolist(),
                [
                    "KXWCGAME-26JUN13-USAMEX-USA",
                    "KXWCROUND-26RO16-USA",
                    "KXWCROUND-26QUAR-USA",
                    "KXWCROUND-26SEMI-USA",
                    "KXMENWORLDCUP-26-AR",
                ],
            )
            self.assertIn("World Soccer Cup Round of 16 Qualifiers", pulled["game"].tolist())
            self.assertIn("2026 World Soccer Cup Winner", pulled["game"].tolist())
            self.assertNotIn("KXWCROUND-26FINAL-USA", pulled["market_ticker"].tolist())
            self.assertNotIn("KXWCHOST-2038-GER", pulled["market_ticker"].tolist())

    def test_kalshi_pull_retries_timeout_with_same_cursor_page(self):
        import requests
        from kalshi_pull import pull_open_events

        first_page = fake_kalshi_response({
            "events": [
                kalshi_event(
                    "KXWCGAME-26JUN13-USAMEX",
                    "KXWCGAME",
                    "USA vs Mexico",
                    [],
                ),
            ],
            "cursor": "next-page",
        })
        second_page = fake_kalshi_response({
            "events": [
                kalshi_event(
                    "KXMENWORLDCUP-26",
                    "KXMENWORLDCUP",
                    "2026 World Soccer Cup Winner",
                    [],
                ),
            ],
            "cursor": None,
        })

        with patch("kalshi_pull.requests.get") as get, patch("kalshi_pull.time.sleep") as sleep:
            get.side_effect = [
                first_page,
                requests.exceptions.Timeout("slow page"),
                second_page,
            ]

            events = pull_open_events()

        self.assertEqual(len(events), 2)
        self.assertEqual(sleep.call_args_list[0].args[0], 2)
        self.assertEqual(get.call_args_list[0].kwargs["params"].get("cursor"), None)
        self.assertEqual(get.call_args_list[1].kwargs["params"].get("cursor"), "next-page")
        self.assertEqual(get.call_args_list[2].kwargs["params"].get("cursor"), "next-page")
        self.assertEqual(get.call_args_list[1].kwargs["timeout"], 45)

    def test_kalshi_pull_retries_retryable_http_status(self):
        from kalshi_pull import get_with_retries

        with patch("kalshi_pull.requests.get") as get, patch("kalshi_pull.time.sleep") as sleep:
            get.side_effect = [
                fake_kalshi_response({"events": []}, status_code=503),
                fake_kalshi_response({"events": []}, status_code=200),
            ]

            response = get_with_retries("https://example.test/events", params={"cursor": "abc"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(get.call_count, 2)
        self.assertEqual(sleep.call_args_list[0].args[0], 2)

    def test_kalshi_pull_retries_chunked_encoding_error_with_same_cursor_page(self):
        import requests
        from kalshi_pull import pull_open_events

        first_page = fake_kalshi_response({
            "events": [
                kalshi_event(
                    "KXWCGAME-26JUN13-USAMEX",
                    "KXWCGAME",
                    "USA vs Mexico",
                    [],
                ),
            ],
            "cursor": "chunk-retry-page",
        })
        second_page = fake_kalshi_response({
            "events": [
                kalshi_event(
                    "KXWCROUND-26RO16",
                    "KXWCROUND",
                    "World Soccer Cup Round of 16 Qualifiers",
                    [],
                ),
            ],
            "cursor": None,
        })

        with patch("kalshi_pull.requests.get") as get, patch("kalshi_pull.time.sleep") as sleep:
            get.side_effect = [
                first_page,
                requests.exceptions.ChunkedEncodingError("incomplete read"),
                second_page,
            ]

            events = pull_open_events()

        self.assertEqual(len(events), 2)
        self.assertEqual(sleep.call_args_list[0].args[0], 2)
        self.assertEqual(get.call_args_list[1].kwargs["params"].get("cursor"), "chunk-retry-page")
        self.assertEqual(get.call_args_list[2].kwargs["params"].get("cursor"), "chunk-retry-page")

    def test_build_value_board_uses_sample_silver_kalshi_and_active_wager_files(self):
        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)
            write_sample_silver_workbook(workspace / "Silver_Current.xlsx")
            write_sample_kalshi_workbook(workspace / "Kalshi_Current.xlsx")
            write_sample_wager_log(workspace / "World_Cup_Bet_Log.xlsx")

            with changed_dir(workspace):
                run_pipeline_script(ROOT / "build_value_board.py")

            output = workspace / "WorldCup_ValueBoard.xlsx"
            self.assertTrue(output.exists())

            bet_sheet = pd.read_excel(output, sheet_name="Bet Sheet")
            metadata = pd.read_excel(output, sheet_name="Run Metadata")
            value_board = pd.read_excel(output, sheet_name="Value Board")
            portfolio_sections = read_portfolio_sections(output)

            self.assertEqual(len(value_board), 3)
            self.assertEqual(metadata.loc[metadata["Field"] == "QC Status", "Value"].iloc[0], "PASS")

            usa = bet_sheet[bet_sheet["market_ticker"] == "KXWCGAME-26JUN13-USAMEX-USA"].iloc[0]
            self.assertEqual(usa["Action"], "BUY YES")
            self.assertEqual(usa["Bucket"], "A")
            self.assertAlmostEqual(usa["Edge"], 0.11)
            self.assertAlmostEqual(usa["Half Kelly"], 0.11)
            self.assertAlmostEqual(usa["Stake on $500"], 53.92)
            self.assertNotIn("Quarter Kelly", bet_sheet.columns)
            self.assertEqual(usa["Position"], "Yes")
            self.assertEqual(usa["Current Action"], "BUY YES")
            self.assertAlmostEqual(usa["Entry"], 0.45)
            self.assertAlmostEqual(usa["Current Value Change"], 0.04)

            mex = bet_sheet[bet_sheet["market_ticker"] == "KXWCGAME-26JUN13-USAMEX-MEX"].iloc[0]
            self.assertEqual(mex["Action"], "SELL YES")
            self.assertEqual(mex["Position"], "No")

            self.assertEqual(
                set(portfolio_sections),
                {
                    "Portfolio Summary",
                    "Open Positions",
                    "Exposure by Team",
                    "Exposure by Match Date",
                },
            )

            portfolio_summary = frame_for_section(portfolio_sections["Portfolio Summary"])
            open_positions = frame_for_section(portfolio_sections["Open Positions"])
            exposure_by_team = frame_for_section(portfolio_sections["Exposure by Team"])
            exposure_by_match_date = frame_for_section(portfolio_sections["Exposure by Match Date"])

            self.assertEqual(portfolio_summary.loc["Open Positions", "Value"], 1)
            self.assertAlmostEqual(portfolio_summary.loc["Total Stake", "Value"], 9.00)
            self.assertAlmostEqual(portfolio_summary.loc["Current Value", "Value"], 9.60)
            self.assertAlmostEqual(portfolio_summary.loc["Unrealized P/L", "Value"], 0.60)

            open_positions = open_positions.reset_index().set_index("Market Ticker")
            usa_position = open_positions.loc["KXWCGAME-26JUN13-USAMEX-USA"]
            self.assertEqual(usa_position["Match Date"], "2026-06-13")
            self.assertEqual(usa_position["Match"], "USA vs Mexico")
            self.assertEqual(usa_position["Outcome"], "United States")
            self.assertEqual(usa_position["Team"], "USA")
            self.assertEqual(usa_position["Position"], "Yes")
            self.assertAlmostEqual(usa_position["Current Price"], 0.48)
            self.assertAlmostEqual(usa_position["Stake"], 9.00)
            self.assertAlmostEqual(usa_position["Current Value"], 9.60)
            self.assertAlmostEqual(usa_position["Unrealized P/L"], 0.60)

            self.assertAlmostEqual(exposure_by_team.loc["USA", "Stake"], 9.00)
            self.assertAlmostEqual(exposure_by_match_date.loc["2026-06-13", "Current Value"], 9.60)

    def test_ladder_math_yes_position(self):
        from build_ladder_board import calculate_ladder

        ladder = calculate_ladder(entry_price=0.50, fair_value=0.74, contracts_held=148)

        self.assertAlmostEqual(ladder["current_edge"], 0.24)
        self.assertEqual(ladder["ladder_1_price"], 0.62)
        self.assertEqual(ladder["ladder_2_price"], 0.74)
        self.assertEqual(ladder["ladder_3_price"], 0.89)
        self.assertEqual(ladder["ladder_1_contracts"], 44)
        self.assertEqual(ladder["ladder_2_contracts"], 44)
        self.assertEqual(ladder["ladder_3_contracts"], 60)

    def test_ladder_math_no_position_uses_no_probability_and_price(self):
        from build_ladder_board import build_ladder_board

        active_wagers = pd.DataFrame([
            {
                "Market Ticker": "KXWCGAME-26JUN13-USAMEX-MEX",
                "Action": "SELL YES",
                "Entry Price": 0.60,
                "Contracts": 10,
                "Status": "Partially Closed",
            }
        ])
        value_rows = pd.DataFrame([
            {
                "market_ticker": "KXWCGAME-26JUN13-USAMEX-MEX",
                "event_title": "USA vs Mexico",
                "market_title": "Mexico wins",
                "yes_bid": 0.61,
                "yes_ask": 0.62,
                "silver_probability_yes": 0.25,
                "market_type": "Match",
            }
        ])

        board = build_ladder_board(active_wagers, value_rows)
        row = board.iloc[0]

        self.assertEqual(row["side"], "NO")
        self.assertAlmostEqual(row["avg_price"], 0.40)
        self.assertAlmostEqual(row["current_price"], 0.38)
        self.assertAlmostEqual(row["silver_probability"], 0.75)
        self.assertAlmostEqual(row["current_edge"], 0.35)
        self.assertEqual(row["ladder_1_price"], 0.58)
        self.assertEqual(row["ladder_2_price"], 0.75)
        self.assertEqual(row["ladder_3_price"], 0.90)

    def test_ladder_contract_allocation_caps_and_no_edge_display(self):
        from build_ladder_board import allocate_contracts, calculate_ladder

        self.assertEqual(allocate_contracts(2), (0, 0, 2))
        self.assertEqual(allocate_contracts(10), (3, 3, 4))

        capped = calculate_ladder(entry_price=0.80, fair_value=0.92, contracts_held=5)
        self.assertEqual(capped["ladder_3_price"], 0.99)

        no_ladder = calculate_ladder(entry_price=0.60, fair_value=0.55, contracts_held=5)
        self.assertAlmostEqual(no_ladder["current_edge"], -0.05)
        self.assertEqual(no_ladder["ladder_display"], "")
        self.assertTrue(pd.isna(no_ladder["ladder_1_price"]))

        display = calculate_ladder(entry_price=0.50, fair_value=0.74, contracts_held=148)
        self.assertEqual(display["ladder_display"], "62¢ × 44 | 74¢ × 44 | 89¢ × 60")

    def test_build_ladder_board_outputs_active_positions_only_and_web_html(self):
        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)
            write_sample_ladder_value_board(workspace / "WorldCup_ValueBoard.xlsx")
            write_sample_ladder_wager_log(workspace / "World_Cup_Bet_Log.xlsx")

            with changed_dir(workspace):
                run_pipeline_script(ROOT / "build_ladder_board.py")
                run_pipeline_script(ROOT / "build_web_ladder_board.py")

            output = workspace / "WorldCup_Ladder_Board.xlsx"
            html_output = workspace / "WorldCup_Ladder_Board.html"
            self.assertTrue(output.exists())
            self.assertTrue(html_output.exists())

            board = pd.read_excel(output, sheet_name="Ladder Board")
            self.assertEqual(set(board["market_ticker"]), {
                "KXWCGAME-26JUN13-USAMEX-USA",
                "KXWCGAME-26JUN13-USAMEX-MEX",
            })
            self.assertNotIn("KXWCGAME-26JUN13-USAMEX-TIE", board["market_ticker"].tolist())

            yes = board[board["market_ticker"] == "KXWCGAME-26JUN13-USAMEX-USA"].iloc[0]
            self.assertEqual(yes["side"], "YES")
            self.assertEqual(yes["ladder_display"], "53¢ × 6 | 60¢ × 6 | 75¢ × 8")
            self.assertEqual(yes["action_flag"], "near")

            no = board[board["market_ticker"] == "KXWCGAME-26JUN13-USAMEX-MEX"].iloc[0]
            self.assertEqual(no["side"], "NO")
            self.assertAlmostEqual(no["avg_price"], 0.40)
            self.assertAlmostEqual(no["silver_probability"], 0.75)

            html = html_output.read_text(encoding="utf-8")
            self.assertIn("World Cup Ladder Board", html)
            self.assertIn("53¢ × 6 | 60¢ × 6 | 75¢ × 8", html)
            self.assertIn('class="near"', html)

    def test_ladder_board_excludes_zero_remaining_contracts_even_if_partially_closed(self):
        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)
            write_sample_ladder_value_board(workspace / "WorldCup_ValueBoard.xlsx")
            write_zero_remaining_ladder_wager_log(workspace / "World_Cup_Bet_Log.xlsx")

            with changed_dir(workspace):
                run_pipeline_script(ROOT / "build_ladder_board.py")

            board = pd.read_excel(workspace / "WorldCup_Ladder_Board.xlsx", sheet_name="Ladder Board")
            self.assertEqual(board["market_ticker"].tolist(), ["KXWCGAME-26JUN13-USAMEX-USA"])
            usa = board.iloc[0]
            self.assertEqual(usa["contracts_held"], 5)
            self.assertEqual(usa["ladder_display"], "53¢ × 1 | 60¢ × 1 | 75¢ × 3")

    def test_ladder_board_excludes_inactive_status_even_with_positive_contracts(self):
        from build_ladder_board import read_active_wagers

        with TemporaryDirectory() as workspace:
            path = Path(workspace) / "World_Cup_Bet_Log.xlsx"
            wagers = pd.DataFrame([
                {
                    "Market Ticker": "OPEN",
                    "Action": "BUY YES",
                    "Entry Price": 0.40,
                    "Contracts": 8,
                    "Status": "Open",
                },
                {
                    "Market Ticker": "SETTLED-LEGACY",
                    "Action": "BUY YES",
                    "Entry Price": 0.40,
                    "Contracts": 8,
                    "Status": "Settled",
                },
                {
                    "Market Ticker": "CLOSED-LEGACY",
                    "Action": "SELL YES",
                    "Entry Price": 0.60,
                    "Contracts": 8,
                    "Remaining Contracts": 8,
                    "Status": "Closed Early",
                },
            ])
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                wagers.to_excel(writer, sheet_name="Wagers", index=False)

            active = read_active_wagers(path)

        self.assertEqual(active["Market Ticker"].tolist(), ["OPEN"])
        self.assertEqual(active["Contracts"].tolist(), [8])

    def test_ladder_board_order_is_deterministic_for_equal_priority_rows(self):
        from build_ladder_board import build_ladder_board

        active_wagers = pd.DataFrame([
            {
                "Market Ticker": ticker,
                "Action": "BUY YES",
                "Entry Price": 0.40,
                "Contracts": 10,
                "Status": "Open",
            }
            for ticker in ["TICKER-B", "TICKER-A"]
        ])
        value_rows = pd.DataFrame([
            {
                "market_ticker": ticker,
                "event_title": "Same event",
                "market_title": "Same market",
                "yes_bid": 0.40,
                "yes_ask": 0.41,
                "silver_probability_yes": 0.60,
                "market_type": "Match",
            }
            for ticker in ["TICKER-B", "TICKER-A"]
        ])

        board = build_ladder_board(active_wagers, value_rows)

        self.assertEqual(board["market_ticker"].tolist(), ["TICKER-A", "TICKER-B"])
        for _, row in board.iterrows():
            allocated = (
                row["ladder_1_contracts"]
                + row["ladder_2_contracts"]
                + row["ladder_3_contracts"]
            )
            self.assertEqual(allocated, row["contracts_held"])

    def test_web_ladder_board_renders_empty_results(self):
        from build_ladder_board import OUTPUT_COLUMNS
        from build_web_ladder_board import render_html

        html = render_html(
            pd.DataFrame(columns=OUTPUT_COLUMNS),
            generated_at=pd.Timestamp("2026-07-28 12:00"),
        )

        self.assertIn("World Cup Ladder Board", html)
        self.assertIn("Generated 2026-07-28 12:00", html)
        self.assertIn("<tbody>", html)
        self.assertNotIn("<tbody>\\n                <tr", html)

    def test_ladder_value_board_reader_deduplicates_columns_before_concat(self):
        from build_ladder_board import (
            VALUE_ROW_COLUMNS,
            make_unique_columns,
            read_value_board_rows,
        )

        self.assertEqual(
            make_unique_columns(["edge", "edge", "edge"]),
            ["edge", "edge__2", "edge__3"],
        )

        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)
            write_duplicate_header_ladder_value_board(workspace / "WorldCup_ValueBoard.xlsx")
            write_duplicate_header_ladder_value_board(workspace / "WorldCup_Futures_ValueBoard.xlsx")
            write_sample_ladder_wager_log(workspace / "World_Cup_Bet_Log.xlsx")

            with changed_dir(workspace):
                match_rows = read_value_board_rows("WorldCup_ValueBoard.xlsx", "Match")
                futures_rows = read_value_board_rows("WorldCup_Futures_ValueBoard.xlsx", "Futures")
                self.assertTrue(match_rows.columns.is_unique)
                self.assertTrue(futures_rows.columns.is_unique)
                self.assertEqual(match_rows.columns.tolist(), VALUE_ROW_COLUMNS)
                pd.concat([match_rows, futures_rows], ignore_index=True)
                run_pipeline_script(ROOT / "build_ladder_board.py")

            output = workspace / "WorldCup_Ladder_Board.xlsx"
            self.assertTrue(output.exists())
            board = pd.read_excel(output, sheet_name="Ladder Board")
            self.assertTrue(board.columns.is_unique)
            self.assertIn("KXWCGAME-26JUN13-USAMEX-USA", board["market_ticker"].tolist())

    def test_build_futures_value_board_matches_champion_and_advancement_markets(self):
        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)
            write_sample_silver_futures_workbook(workspace / "Silver_Futures_Current.xlsx")
            write_sample_kalshi_futures_workbook(workspace / "Kalshi_Current.xlsx")
            write_sample_futures_wager_log(workspace / "World_Cup_Bet_Log.xlsx")

            with changed_dir(workspace):
                run_pipeline_script(ROOT / "build_futures_value_board.py")

            output = workspace / "WorldCup_Futures_ValueBoard.xlsx"
            self.assertTrue(output.exists())

            bet_sheet = pd.read_excel(output, sheet_name="Bet Sheet")
            metadata = pd.read_excel(output, sheet_name="Run Metadata")
            value_board = pd.read_excel(output, sheet_name="Value Board")

            self.assertEqual(metadata.loc[metadata["Field"] == "QC Status", "Value"].iloc[0], "PASS")
            self.assertEqual(len(value_board), 35)

            argentina = bet_sheet[bet_sheet["market_ticker"] == "KXMENWORLDCUP-26-AR"].iloc[0]
            self.assertEqual(argentina["Stage"], "Champion")
            self.assertEqual(argentina["Team"], "ARG")
            self.assertEqual(argentina["Action"], "BUY YES")
            self.assertAlmostEqual(argentina["Silver"], 0.241)
            self.assertAlmostEqual(argentina["Market Price"], 0.151)
            self.assertAlmostEqual(argentina["Edge"], 0.09)
            self.assertAlmostEqual(argentina["Quarter Kelly"], 0.03)
            expected_argentina_proxy_kelly = (
                value_board.loc[
                    (value_board["Team"] == "ARG")
                    & (value_board["stage_key"] == "Champ"),
                    "buy_quarter_kelly",
                ].iloc[0]
                + 0.60
                * value_board.loc[
                    (value_board["Team"] == "ARG")
                    & value_board["stage_key"].isin(["R16", "Qtr", "Semi"]),
                    "buy_quarter_kelly",
                ].sum()
            ) * 500
            self.assertAlmostEqual(
                argentina["proxy_kelly_dollars"],
                expected_argentina_proxy_kelly,
                places=2,
            )
            self.assertAlmostEqual(argentina["proxy_kelly_dollars"], 97.45, places=2)
            self.assertEqual(argentina["proxy_contracts"], 645)
            self.assertEqual(
                [
                    argentina["ladder_1_price"],
                    argentina["ladder_2_price"],
                    argentina["ladder_3_price"],
                    argentina["ladder_4_price"],
                ],
                [0.23, 0.34, 0.49, 0.68],
            )
            argentina_ladder_contracts = [
                argentina["ladder_1_contracts"],
                argentina["ladder_2_contracts"],
                argentina["ladder_3_contracts"],
                argentina["ladder_4_contracts"],
            ]
            self.assertEqual(argentina_ladder_contracts, [193, 193, 129, 130])
            self.assertEqual(sum(argentina_ladder_contracts), argentina["proxy_contracts"])
            self.assertAlmostEqual(argentina["ladder_expected_return"], 261.62)
            self.assertAlmostEqual(
                argentina["ladder_expected_profit"],
                argentina["ladder_expected_return"] - argentina["proxy_kelly_dollars"],
                places=2,
            )
            self.assertAlmostEqual(argentina["Stake on $500"], 13.25)
            self.assertEqual(argentina["Bucket"], "B")
            self.assertEqual(argentina["Position"], "Yes")
            self.assertEqual(argentina["Current Action"], "BUY YES")
            self.assertAlmostEqual(argentina["Entry Price"], 0.15)
            self.assertAlmostEqual(argentina["Stake"], 1.80)
            self.assertAlmostEqual(argentina["Current Value Change"], 0.001)
            self.assertEqual(argentina["Status"], "Open")
            self.assertEqual(argentina["champion_proxy_candidate"], "YES")
            self.assertEqual(argentina["strong_champion_proxy"], "YES")

            usa = bet_sheet[bet_sheet["market_ticker"] == "KXWCROUND-26RO16-USA"].iloc[0]
            self.assertEqual(usa["Stage"], "Round of 16")
            self.assertEqual(usa["Action"], "BUY YES")
            self.assertAlmostEqual(usa["Silver"], 0.72)
            self.assertAlmostEqual(usa["Market Price"], 0.61)
            self.assertAlmostEqual(usa["Edge"], 0.11)
            self.assertAlmostEqual(usa["Quarter Kelly"], 0.07)
            self.assertAlmostEqual(usa["Stake on $500"], 35.26)
            self.assertEqual(usa["Position"], "Yes")
            self.assertEqual(usa["Current Action"], "SELL YES")
            self.assertAlmostEqual(usa["Entry Price"], 0.65)
            self.assertAlmostEqual(usa["Stake"], 3.50)
            self.assertAlmostEqual(usa["Current Value Change"], 0.04)
            self.assertEqual(usa["Status"], "Partially Closed")

            quarterfinal = bet_sheet[
                bet_sheet["market_ticker"] == "KXWCROUND-26QUAR-USA"
            ].iloc[0]
            self.assertEqual(quarterfinal["Position"], "No")
            self.assertTrue(pd.isna(quarterfinal["Current Action"]))

            self.assertNotIn("Half Kelly", bet_sheet.columns)
            self.assertIn("event_ticker", bet_sheet.columns)

            norway = value_board[value_board["Team"] == "NOR"]
            norway_value = norway.set_index("stage_key")
            self.assertEqual(set(norway["stage_key"]), {"R16", "Qtr", "Semi", "Champ"})
            self.assertEqual(set(norway["champion_proxy_candidate"]), {"YES"})
            self.assertTrue(norway["strong_champion_proxy"].fillna("").eq("").all())
            expected_norway_proxy_kelly = (
                norway_value.loc["Champ", "buy_quarter_kelly"]
                + 0.60
                * norway_value.loc[["R16", "Qtr", "Semi"], "buy_quarter_kelly"].sum()
            ) * 500
            self.assertEqual(
                set(bet_sheet.loc[bet_sheet["Team"] == "NOR", "proxy_kelly_dollars"].round(2)),
                {round(expected_norway_proxy_kelly, 2)},
            )
            self.assertEqual(
                set(bet_sheet.loc[bet_sheet["Team"] == "NOR", "champion_proxy_candidate"]),
                {"YES"},
            )
            self.assertTrue(
                bet_sheet.loc[bet_sheet["Team"] == "NOR", "strong_champion_proxy"]
                .fillna("")
                .eq("")
                .all()
            )

            brazil = value_board[value_board["Team"] == "BRA"]
            self.assertLess(
                brazil.loc[brazil["stage_key"] == "Champ", "buy_edge"].iloc[0],
                0,
            )
            self.assertTrue(
                brazil["champion_proxy_candidate"].fillna("").eq("").all()
            )
            self.assertTrue(brazil["strong_champion_proxy"].fillna("").eq("").all())
            self.assertTrue(brazil["proxy_kelly_dollars"].fillna("").eq("").all())
            self.assertTrue(brazil["proxy_contracts"].fillna("").eq("").all())
            self.assertTrue(brazil["ladder_expected_profit"].fillna("").eq("").all())

            norway_stage_average = norway_value.loc[["R16", "Qtr", "Semi"], "buy_edge"].mean()
            self.assertGreater(norway_value.loc["Champ", "buy_edge"], 0)
            self.assertLess(
                norway_value.loc["Champ", "buy_edge"],
                0.35 * norway_stage_average,
            )

            france = value_board[value_board["Team"] == "FRA"]
            self.assertNotIn("Semi", set(france["stage_key"]))
            self.assertTrue(france["champion_proxy_candidate"].fillna("").eq("").all())
            self.assertTrue(france["strong_champion_proxy"].fillna("").eq("").all())
            self.assertTrue(france["proxy_kelly_dollars"].fillna("").eq("").all())
            self.assertTrue(france["proxy_contracts"].fillna("").eq("").all())

            mexico = value_board[value_board["Team"] == "MEX"]
            self.assertTrue(mexico["buy_quarter_kelly"].isna().any())
            self.assertTrue(mexico["proxy_kelly_dollars"].fillna("").eq("").all())
            self.assertTrue(mexico["proxy_contracts"].fillna("").eq("").all())

            australia = value_board[value_board["Team"] == "AUS"].iloc[0]
            self.assertEqual(australia["champion_proxy_candidate"], "YES")
            self.assertEqual(australia["proxy_contracts"], 438)
            self.assertEqual(
                [
                    australia["ladder_1_price"],
                    australia["ladder_2_price"],
                    australia["ladder_3_price"],
                    australia["ladder_4_price"],
                ],
                [0.12, 0.18, 0.26, 0.36],
            )

            england = value_board[value_board["Team"] == "ENG"].iloc[0]
            self.assertEqual(england["champion_proxy_candidate"], "YES")
            self.assertEqual(england["ladder_4_price"], 0.99)

            germany = value_board[value_board["Team"] == "GER"]
            self.assertEqual(set(germany["champion_proxy_candidate"]), {"YES"})
            self.assertTrue(germany["proxy_kelly_dollars"].fillna("").ne("").all())
            self.assertTrue(germany["proxy_contracts"].fillna("").eq("").all())

            workbook = load_workbook(output, data_only=False)
            bet_sheet_ws = workbook["Bet Sheet"]
            bet_headers = {cell.value: cell.column for cell in bet_sheet_ws[1]}
            self.assertEqual(bet_sheet_ws.cell(2, bet_headers["Silver"]).number_format, "0.0%")
            self.assertEqual(
                bet_sheet_ws.cell(2, bet_headers["Edge"]).number_format,
                "+0.0%;-0.0%;0.0%",
            )
            self.assertIsInstance(bet_sheet_ws.cell(2, bet_headers["Silver"]).value, float)
            self.assertEqual(
                bet_sheet_ws.cell(2, bet_headers["proxy_kelly_dollars"]).number_format,
                "$0.00",
            )
            self.assertEqual(
                bet_sheet_ws.cell(2, bet_headers["ladder_1_price"]).number_format,
                "0%",
            )
            self.assertEqual(
                bet_sheet_ws.cell(2, bet_headers["proxy_contracts"]).number_format,
                "0",
            )
            self.assertEqual(
                bet_sheet_ws.cell(2, bet_headers["ladder_expected_profit"]).number_format,
                "$0.00",
            )

    def test_build_web_futures_betsheet_creates_sortable_html_with_expected_rows(self):
        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)
            write_sample_futures_value_board(workspace / "WorldCup_Futures_ValueBoard.xlsx")

            with changed_dir(workspace):
                run_pipeline_script(ROOT / "build_web_futures_betsheet.py")

            output = workspace / "WorldCup_Futures_ValueBoard.html"
            self.assertTrue(output.exists())

            html = output.read_text(encoding="utf-8")
            self.assertIn("World Cup Futures Value Board", html)
            self.assertIn("sortTable(table, columnIndex, ascending)", html)
            self.assertIn("Champion", html)
            self.assertIn("ARG", html)
            self.assertIn("KXMENWORLDCUP-26-AR", html)
            self.assertIn("<th>Quarter Kelly</th>", html)
            self.assertIn("<th>proxy_kelly_dollars</th>", html)
            self.assertIn("<th>proxy_contracts</th>", html)
            self.assertIn("<th>proxy_sell_ladder</th>", html)
            self.assertIn("<th>ladder_expected_return</th>", html)
            self.assertIn("<th>ladder_expected_profit</th>", html)
            self.assertIn("<th>Position</th>", html)
            self.assertIn("<th>Current Action</th>", html)
            self.assertIn("<th>Entry Price</th>", html)
            self.assertIn("<th>Current Value Change</th>", html)
            self.assertIn("<th>champion_proxy_candidate</th>", html)
            self.assertIn("<th>strong_champion_proxy</th>", html)
            self.assertIn("Champion Proxy Candidate means R16/QF/SF/Champion", html)
            self.assertIn("Proxy Kelly is shown only for Champion Proxy teams", html)
            self.assertIn("Champion Proxy sell ladders use the Champion BUY price", html)
            self.assertIn('["champion_proxy_candidate", "proxy-badge"]', html)
            self.assertIn('["strong_champion_proxy", "strong-proxy-badge"]', html)
            self.assertIn("24.1%", html)
            self.assertIn("+9.1%", html)
            self.assertIn("$50.33", html)
            self.assertIn("23¢ × 193 | 34¢ × 193 | 49¢ × 129 | 68¢ × 130", html)
            self.assertIn("$261.62", html)
            self.assertIn("$164.17", html)
            self.assertNotIn("<th>event_ticker</th>", html)
            self.assertNotIn("Half Kelly", html)
            self.assertNotIn(">nan<", html)
            self.assertNotIn("NaN", html)

    def test_import_wagers_uses_sample_activity_and_value_board_lookup(self):
        with TemporaryDirectory() as home:
            project_dir = Path(home) / "kalshi"
            project_dir.mkdir()
            shutil.copy2(
                FIXTURES / "sample_kalshi_activity.csv",
                project_dir / "Kalshi-Recent-Activity-All-sample.csv",
            )
            write_sample_value_board_lookup(project_dir / "WorldCup_ValueBoard.xlsx")

            with patched_environ(HOME=home):
                run_pipeline_script(ROOT / "import_wagers.py")

            output = project_dir / "World_Cup_Bet_Log.xlsx"
            self.assertTrue(output.exists())

            wagers = pd.read_excel(output, sheet_name="Wagers")
            summary = pd.read_excel(output, sheet_name="Summary")

            self.assertEqual(len(wagers), 3)
            self.assertEqual(summary.loc[summary["Metric"] == "Total Wagers", "Value"].iloc[0], 3)
            self.assertEqual(len(wager_log_archives(project_dir)), 1)

            usa = wagers[wagers["Market Ticker"] == "KXWCGAME-26JUN13-USAMEX-USA"].iloc[0]
            self.assertEqual(usa["Market Type"], "Match")
            self.assertTrue(pd.isna(usa["Stage"]))
            self.assertEqual(usa["Team"], "USA")
            self.assertEqual(usa["Action"], "BUY YES")
            self.assertEqual(usa["Status"], "Open")
            self.assertEqual(usa["Match Date"], "2026-06-13")
            self.assertEqual(usa["Match"], "USA vs Mexico")
            self.assertEqual(usa["Outcome"], "United States")
            self.assertAlmostEqual(usa["Entry Price"], 0.45)
            self.assertAlmostEqual(usa["Exit Contracts"], 0)
            self.assertAlmostEqual(usa["Remaining Contracts"], usa["Contracts"])

            mex = wagers[wagers["Market Ticker"] == "KXWCGAME-26JUN13-USAMEX-MEX"].iloc[0]
            self.assertEqual(mex["Action"], "SELL YES")
            self.assertEqual(mex["Status"], "Partially Closed")
            self.assertAlmostEqual(mex["Exit Price"], 0.30)
            self.assertGreater(mex["Remaining Contracts"], 0)
            self.assertLess(mex["Remaining Contracts"], mex["Contracts"])

            bra = wagers[wagers["Market Ticker"] == "KXWCGAME-26JUN14-CANBRA-BRA"].iloc[0]
            self.assertEqual(bra["Status"], "Settled")
            self.assertEqual(bra["Contract Won"], "Yes")
            self.assertAlmostEqual(bra["Realized P/L"], 2.20)
            self.assertAlmostEqual(bra["Remaining Contracts"], 0)

    def test_import_wagers_enriches_round_and_champion_futures_wagers(self):
        with TemporaryDirectory() as home:
            project_dir = Path(home) / "kalshi"
            project_dir.mkdir()

            activity = pd.DataFrame(
                [
                    trade("2026-06-20 09:00:00", "Yes", "KXWCROUND-26RO16-USA", 10, 61, 0.10),
                    settlement("2026-07-02 18:00:00", "KXWCROUND-26RO16-USA", "yes"),
                    trade("2026-06-20 10:00:00", "Yes", "KXMENWORLDCUP-26-AR", 12, 15, 0.12),
                ]
            )
            activity.to_csv(project_dir / "Kalshi-Recent-Activity-All-futures.csv", index=False)
            write_sample_futures_value_board(project_dir / "WorldCup_Futures_ValueBoard.xlsx")
            write_sample_kalshi_futures_workbook(project_dir / "Kalshi_Current.xlsx")

            with patched_environ(HOME=home):
                run_pipeline_script(ROOT / "import_wagers.py")

            wagers = pd.read_excel(project_dir / "World_Cup_Bet_Log.xlsx", sheet_name="Wagers")
            self.assertEqual(len(wagers), 2)

            round_of_16 = row_for(wagers, "KXWCROUND-26RO16-USA")
            self.assertEqual(round_of_16["Market Type"], "Futures")
            self.assertEqual(round_of_16["Stage"], "Round of 16")
            self.assertEqual(round_of_16["Team"], "USA")
            self.assertEqual(round_of_16["Outcome"], "United States")
            self.assertTrue(pd.isna(round_of_16["Match Date"]))
            self.assertTrue(pd.isna(round_of_16["Match"]))
            self.assertAlmostEqual(round_of_16["Silver Probability"], 0.72)
            self.assertAlmostEqual(round_of_16["Edge"], 0.11)
            self.assertEqual(round_of_16["Bucket"], "A")
            self.assertEqual(round_of_16["Status"], "Settled")
            self.assertEqual(round_of_16["Contract Won"], "Yes")
            self.assertAlmostEqual(round_of_16["Realized P/L"], 3.80)

            champion = row_for(wagers, "KXMENWORLDCUP-26-AR")
            self.assertEqual(champion["Market Type"], "Futures")
            self.assertEqual(champion["Stage"], "Champion")
            self.assertEqual(champion["Team"], "ARG")
            self.assertEqual(champion["Outcome"], "Argentina")
            self.assertTrue(pd.isna(champion["Match Date"]))
            self.assertTrue(pd.isna(champion["Match"]))
            self.assertAlmostEqual(champion["Silver Probability"], 0.241)
            self.assertAlmostEqual(champion["Edge"], 0.091)
            self.assertEqual(champion["Bucket"], "B")
            self.assertEqual(champion["Status"], "Open")

    def test_import_wagers_covers_exits_settlements_multiple_fills_and_fallbacks(self):
        with TemporaryDirectory() as home:
            project_dir = Path(home) / "kalshi"
            project_dir.mkdir()

            activity = pd.DataFrame(
                [
                    trade("2026-06-01 09:00:00", "Yes", "KXWCGAME-26JUL01-AAAEEE-AAA", 10, 40, 0.10),
                    trade("2026-06-01 10:00:00", "No", "KXWCGAME-26JUL01-AAAEEE-AAA", 10, 55, 0.05),
                    trade("2026-06-02 09:00:00", "Yes", "KXWCGAME-26JUL02-BBBFFF-BBB", 8, 35, 0.08),
                    settlement("2026-07-03 18:00:00", "KXWCGAME-26JUL02-BBBFFF-BBB", "no"),
                    trade("2026-06-03 09:00:00", "No", "KXWCGAME-26JUL03-CCCGGG-CCC", 12, 62, 0.12),
                    settlement("2026-07-04 18:00:00", "KXWCGAME-26JUL03-CCCGGG-CCC", "no"),
                    trade("2026-06-04 09:00:00", "Yes", "KXWCGAME-26JUL04-DDDHHH-DDD", 10, 40, 0.10),
                    trade("2026-06-04 09:05:00", "Yes", "KXWCGAME-26JUL04-DDDHHH-DDD", 30, 50, 0.30),
                    trade("2026-06-05 09:00:00", "Yes", "KXWCGAME-26AUG05-IIIJJJ-III", 5, 25, 0.05),
                    trade("2026-06-06 09:00:00", "Yes", "KXWCGAME-26SEP06-KKKLLL-KKK", 7, 45, 0.07),
                ]
            )
            activity.to_csv(project_dir / "Kalshi-Recent-Activity-All-expanded.csv", index=False)

            write_expanded_value_board_lookup(project_dir / "WorldCup_ValueBoard.xlsx")
            write_prior_wager_log_for_preservation(project_dir / "World_Cup_Bet_Log.xlsx")

            with patched_environ(HOME=home):
                run_pipeline_script(ROOT / "import_wagers.py")

            output = project_dir / "World_Cup_Bet_Log.xlsx"
            wagers = pd.read_excel(output, sheet_name="Wagers")
            summary = pd.read_excel(output, sheet_name="Summary")

            self.assertEqual(len(wagers), 6)
            self.assertEqual(summary.loc[summary["Metric"] == "Total Wagers", "Value"].iloc[0], 6)
            self.assertEqual(summary.loc[summary["Metric"] == "Closed Early", "Value"].iloc[0], 1)
            self.assertEqual(summary.loc[summary["Metric"] == "Settled", "Value"].iloc[0], 2)
            self.assertEqual(summary.loc[summary["Metric"] == "Open", "Value"].iloc[0], 3)
            self.assertEqual(len(wager_log_archives(project_dir)), 1)

            closed = row_for(wagers, "KXWCGAME-26JUL01-AAAEEE-AAA")
            self.assertEqual(closed["Match Date"], "2026-07-01")
            self.assertEqual(closed["Action"], "BUY YES")
            self.assertEqual(closed["Status"], "Closed Early")
            self.assertEqual(closed["Market Result"], "yes")
            self.assertEqual(closed["Contract Won"], "Yes")
            self.assertAlmostEqual(closed["Entry Price"], 0.40)
            self.assertAlmostEqual(closed["Exit Price"], 0.55)
            self.assertAlmostEqual(closed["Realized P/L"], 1.35)
            self.assertAlmostEqual(closed["Exit Contracts"], 10)
            self.assertAlmostEqual(closed["Remaining Contracts"], 0)

            losing_buy = row_for(wagers, "KXWCGAME-26JUL02-BBBFFF-BBB")
            self.assertEqual(losing_buy["Match Date"], "2026-07-02")
            self.assertEqual(losing_buy["Action"], "BUY YES")
            self.assertEqual(losing_buy["Status"], "Settled")
            self.assertEqual(losing_buy["Market Result"], "no")
            self.assertEqual(losing_buy["Contract Won"], "No")
            self.assertAlmostEqual(losing_buy["Realized P/L"], -2.88)
            self.assertAlmostEqual(losing_buy["Remaining Contracts"], 0)
            self.assertAlmostEqual(losing_buy["Closing Price"], 0.45)
            self.assertAlmostEqual(losing_buy["CLV"], 0.10)
            self.assertAlmostEqual(losing_buy["CLV %"], 0.2857)

            winning_sell = row_for(wagers, "KXWCGAME-26JUL03-CCCGGG-CCC")
            self.assertEqual(winning_sell["Action"], "SELL YES")
            self.assertEqual(winning_sell["Status"], "Settled")
            self.assertEqual(winning_sell["Market Result"], "no")
            self.assertEqual(winning_sell["Contract Won"], "Yes")
            self.assertAlmostEqual(winning_sell["Realized P/L"], 7.32)
            self.assertAlmostEqual(winning_sell["Remaining Contracts"], 0)
            self.assertAlmostEqual(winning_sell["Closing Price"], 0.50)
            self.assertAlmostEqual(winning_sell["CLV"], 0.12)
            self.assertAlmostEqual(winning_sell["CLV %"], 0.1935)

            multiple_fill = row_for(wagers, "KXWCGAME-26JUL04-DDDHHH-DDD")
            self.assertEqual(multiple_fill["Action"], "BUY YES")
            self.assertEqual(multiple_fill["Status"], "Open")
            self.assertAlmostEqual(multiple_fill["Contracts"], 40)
            self.assertAlmostEqual(multiple_fill["Exit Contracts"], 0)
            self.assertAlmostEqual(multiple_fill["Remaining Contracts"], 40)
            self.assertAlmostEqual(multiple_fill["Entry Price"], 0.475)
            self.assertAlmostEqual(multiple_fill["Fee In"], 0.40)
            self.assertTrue(pd.isna(multiple_fill["Closing Price"]))
            self.assertAlmostEqual(multiple_fill["CLV"], 0.03)
            self.assertAlmostEqual(multiple_fill["CLV %"], 0.0632)

            preserved = row_for(wagers, "KXWCGAME-26AUG05-IIIJJJ-III")
            self.assertEqual(preserved["Match Date"], "2026-08-05")
            self.assertEqual(preserved["Match"], "Prior Match")
            self.assertEqual(preserved["Outcome"], "Prior Outcome")
            self.assertAlmostEqual(preserved["Silver Probability"], 0.42)
            self.assertAlmostEqual(preserved["Edge"], 0.08)
            self.assertEqual(preserved["Bucket"], "B")
            self.assertAlmostEqual(preserved["Closing Price"], 0.64)
            self.assertAlmostEqual(preserved["CLV"], 0.39)
            self.assertAlmostEqual(preserved["CLV %"], 1.56)

            fallback = row_for(wagers, "KXWCGAME-26SEP06-KKKLLL-KKK")
            self.assertEqual(fallback["Match Date"], "2026-09-06")
            self.assertTrue(pd.isna(fallback["Match"]))
            self.assertTrue(pd.isna(fallback["Outcome"]))
            self.assertTrue(pd.isna(fallback["Market Result"]))
            self.assertTrue(pd.isna(fallback["Contract Won"]))
            self.assertTrue(pd.isna(fallback["Closing Price"]))
            self.assertTrue(pd.isna(fallback["CLV"]))
            self.assertTrue(pd.isna(fallback["CLV %"]))

    def test_import_wagers_normalizes_match_dates_from_strings_timestamps_and_tickers(self):
        with TemporaryDirectory() as home:
            project_dir = Path(home) / "kalshi"
            project_dir.mkdir()

            activity = pd.DataFrame(
                [
                    trade("2026-06-01 09:00:00", "Yes", "KXWCGAME-26JUN14-STRONE-STR", 10, 40, 0.10),
                    trade("2026-06-02 09:00:00", "Yes", "KXWCGAME-26JUN14-TIMTWO-TIM", 8, 35, 0.08),
                    trade("2026-06-03 09:00:00", "Yes", "KXWCGAME-26JUN14-FALBAK-FAL", 7, 45, 0.07),
                    trade("2026-06-04 09:00:00", "Yes", "KXWCGAME-26JUN12-BADOLD-BAD", 6, 44, 0.06),
                ]
            )
            activity.to_csv(project_dir / "Kalshi-Recent-Activity-All-dates.csv", index=False)
            write_date_format_value_board_lookup(project_dir / "WorldCup_ValueBoard.xlsx")
            write_prior_timestamp_match_date_log(project_dir / "World_Cup_Bet_Log.xlsx")

            with patched_environ(HOME=home):
                run_pipeline_script(ROOT / "import_wagers.py")

            wagers = pd.read_excel(project_dir / "World_Cup_Bet_Log.xlsx", sheet_name="Wagers")

            string_date = row_for(wagers, "KXWCGAME-26JUN14-STRONE-STR")
            self.assertEqual(string_date["Match Date"], "2026-06-14")

            timestamp_date = row_for(wagers, "KXWCGAME-26JUN14-TIMTWO-TIM")
            self.assertEqual(timestamp_date["Match Date"], "2026-06-14")

            ticker_date = row_for(wagers, "KXWCGAME-26JUN14-FALBAK-FAL")
            self.assertEqual(ticker_date["Match Date"], "2026-06-14")

            invalid_prior_date = row_for(wagers, "KXWCGAME-26JUN12-BADOLD-BAD")
            self.assertEqual(invalid_prior_date["Match Date"], "2026-06-12")

    def test_build_web_betlog_creates_sortable_html_with_expected_rows_and_blanks(self):
        with TemporaryDirectory() as workspace:
            workspace = Path(workspace)
            write_sample_web_wager_log(workspace / "World_Cup_Bet_Log.xlsx")

            with changed_dir(workspace):
                run_pipeline_script(ROOT / "build_web_betlog.py")

            output = workspace / "World_Cup_Bet_Log.html"
            self.assertTrue(output.exists())

            html = output.read_text(encoding="utf-8")
            self.assertIn("World Cup Bet Log", html)
            self.assertIn("sortTable(table, index, ascending)", html)
            self.assertIn("<th>Market Type</th>", html)
            self.assertIn("<th>Stage</th>", html)
            self.assertIn("<th>Team</th>", html)
            self.assertIn("USA vs Mexico", html)
            self.assertIn("Canada vs Brazil", html)
            self.assertIn("BUY YES", html)
            self.assertNotIn("NaN", html)
            self.assertIn('data-sort="2026-06-14"', html)
            self.assertIn('data-sort="2026-06-11 10:00:00"', html)
            self.assertIn('data-sort="0.60">0.60</td>', html)
            self.assertIn('data-sort="0.12">0.12</td>', html)
            self.assertIn('data-sort="2.20">2.20</td>', html)


class changed_dir:
    def __init__(self, path):
        self.path = path
        self.previous = None

    def __enter__(self):
        self.previous = Path.cwd()
        os.chdir(self.path)

    def __exit__(self, exc_type, exc, tb):
        os.chdir(self.previous)


class patched_environ:
    def __init__(self, **updates):
        self.updates = updates
        self.previous = {}

    def __enter__(self):
        for key, value in self.updates.items():
            self.previous[key] = os.environ.get(key)
            os.environ[key] = value

    def __exit__(self, exc_type, exc, tb):
        for key, value in self.previous.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value


def run_pipeline_script(path):
    sys.modules.pop("config", None)
    runpy.run_path(str(path), run_name="__main__")


def write_sample_silver_workbook(path):
    forecasts = pd.DataFrame(
        [
            {
                "Date": "2026-06-13",
                "Team": "USA",
                "Win": 60,
                "GF": 1.8,
                "Opponent": "MEX",
                "OpponentWin": 15,
                "OpponentGF": 0.9,
                "Draw": 25,
                "MostLikelyScore": "1-0",
            }
        ]
    )
    metadata = pd.DataFrame(
        [["Silver Source File", "sample_silver.csv"], ["Silver Rows Output", 1]],
        columns=["Field", "Value"],
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        forecasts.to_excel(writer, sheet_name="Silver Forecasts", index=False)
        metadata.to_excel(writer, sheet_name="Silver Metadata", index=False)


def write_sample_kalshi_workbook(path):
    pd.read_csv(FIXTURES / "sample_kalshi_current.csv").to_excel(path, index=False)


class fake_kalshi_response:
    def __init__(self, payload, status_code=200):
        self.payload = payload
        self.status_code = status_code

    def raise_for_status(self):
        if self.status_code >= 400:
            import requests

            raise requests.exceptions.HTTPError(
                f"HTTP {self.status_code}",
                response=self,
            )
        return None

    def json(self):
        return self.payload


def kalshi_event(event_ticker, series_ticker, title, markets):
    return {
        "event_ticker": event_ticker,
        "series_ticker": series_ticker,
        "title": title,
        "sub_title": "",
        "category": "Sports",
        "markets": markets,
    }


def kalshi_market(ticker, yes_sub_title, title):
    return {
        "ticker": ticker,
        "yes_sub_title": yes_sub_title,
        "title": title,
        "yes_bid_dollars": 0.48,
        "yes_ask_dollars": 0.49,
        "last_price_dollars": 0.49,
        "volume_fp": 100,
        "open_interest_fp": 50,
        "close_time": "2026-07-19T23:59:00Z",
    }


def write_sample_silver_futures_workbook(path):
    futures = pd.DataFrame(
        [
            {"Team": "ARG", "Team Raw": "ARG Argentina", "R16": 0.864, "Qtr": 0.552, "Semi": 0.376, "Final": 0.293, "Champ": 0.241},
            {"Team": "USA", "Team Raw": "USA United States", "R16": 0.72, "Qtr": 0.35, "Semi": 0.18, "Final": 0.08, "Champ": 0.03},
            {"Team": "NOR", "Team Raw": "NOR Norway", "R16": 0.80, "Qtr": 0.60, "Semi": 0.40, "Final": 0.28, "Champ": 0.20},
            {"Team": "BRA", "Team Raw": "BRA Brazil", "R16": 0.75, "Qtr": 0.50, "Semi": 0.30, "Final": 0.18, "Champ": 0.10},
            {"Team": "FRA", "Team Raw": "FRA France", "R16": 0.70, "Qtr": 0.45, "Semi": 0.25, "Final": 0.14, "Champ": 0.08},
            {"Team": "MEX", "Team Raw": "MEX Mexico", "R16": 0.76, "Qtr": 0.48, "Semi": 0.28, "Final": 0.12, "Champ": 0.06},
            {"Team": "AUS", "Team Raw": "AUS Australia", "R16": 0.55, "Qtr": 0.34, "Semi": 0.20, "Final": 0.12, "Champ": 0.11},
            {"Team": "ENG", "Team Raw": "ENG England", "R16": 0.82, "Qtr": 0.62, "Semi": 0.48, "Final": 0.40, "Champ": 0.38},
            {"Team": "GER", "Team Raw": "GER Germany", "R16": 0.78, "Qtr": 0.58, "Semi": 0.44, "Final": 0.32, "Champ": 0.18},
        ]
    )
    metadata = pd.DataFrame(
        [["Silver Futures Source File", "sample_silver_futures.csv"], ["Silver Futures Rows Output", 2]],
        columns=["Field", "Value"],
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        futures.to_excel(writer, sheet_name="Silver Futures", index=False)
        metadata.to_excel(writer, sheet_name="Silver Metadata", index=False)


def write_sample_kalshi_futures_workbook(path):
    rows = pd.DataFrame(
        [
            {
                "event_ticker": "KXMENWORLDCUP-26",
                "game": "World Cup winner",
                "sub_title": "",
                "category": "Sports",
                "market_ticker": "KXMENWORLDCUP-26-AR",
                "outcome": "Argentina",
                "market_title": "Will Argentina win the World Cup?",
                "yes_bid_dollars": 0.14,
                "yes_ask_dollars": 0.151,
                "last_price_dollars": 0.151,
                "volume_fp": 1000,
                "open_interest_fp": 500,
                "close_time": "2026-07-19T23:59:00Z",
            },
            {
                "event_ticker": "KXWCROUND-26RO16",
                "game": "Round of 16 Qualifiers",
                "sub_title": "",
                "category": "Sports",
                "market_ticker": "KXWCROUND-26RO16-USA",
                "outcome": "United States",
                "market_title": "Will USA qualify for the Round of 16?",
                "yes_bid_dollars": 0.60,
                "yes_ask_dollars": 0.61,
                "last_price_dollars": 0.61,
                "volume_fp": 800,
                "open_interest_fp": 300,
                "close_time": "2026-07-01T23:59:00Z",
            },
            {
                "event_ticker": "KXWCROUND-26QUAR",
                "game": "Quarterfinals Qualifiers",
                "sub_title": "",
                "category": "Sports",
                "market_ticker": "KXWCROUND-26QUAR-USA",
                "outcome": "United States",
                "market_title": "Will USA qualify for the Quarterfinals?",
                "yes_bid_dollars": 0.42,
                "yes_ask_dollars": 0.43,
                "last_price_dollars": 0.43,
                "volume_fp": 600,
                "open_interest_fp": 200,
                "close_time": "2026-07-05T23:59:00Z",
            },
            {
                "event_ticker": "KXWCGAME-26JUN13-USAMEX",
                "game": "USA vs Mexico",
                "sub_title": "",
                "category": "Sports",
                "market_ticker": "KXWCGAME-26JUN13-USAMEX-USA",
                "outcome": "United States",
                "market_title": "USA vs Mexico Winner?",
                "yes_bid_dollars": 0.48,
                "yes_ask_dollars": 0.49,
                "last_price_dollars": 0.49,
                "volume_fp": 100,
                "open_interest_fp": 50,
                "close_time": "2026-06-13T23:59:00Z",
            },
        ]
    )
    proxy_rows = [
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-NOR", "Norway", "Round of 16 Qualifiers", 0.69, 0.70),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-NOR", "Norway", "Quarterfinals Qualifiers", 0.49, 0.50),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-NOR", "Norway", "Semifinals Qualifiers", 0.31, 0.32),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-NO", "Norway", "World Cup winner", 0.16, 0.17),
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-ARG", "Argentina", "Round of 16 Qualifiers", 0.63, 0.64),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-ARG", "Argentina", "Quarterfinals Qualifiers", 0.34, 0.35),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-ARG", "Argentina", "Semifinals Qualifiers", 0.22, 0.23),
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-BRA", "Brazil", "Round of 16 Qualifiers", 0.64, 0.65),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-BRA", "Brazil", "Quarterfinals Qualifiers", 0.41, 0.42),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-BRA", "Brazil", "Semifinals Qualifiers", 0.21, 0.22),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-BR", "Brazil", "World Cup winner", 0.10, 0.11),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-USA", "United States", "Semifinals Qualifiers", 0.11, 0.12),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-US", "United States", "World Cup winner", 0.01, 0.02),
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-FRA", "France", "Round of 16 Qualifiers", 0.59, 0.60),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-FRA", "France", "Quarterfinals Qualifiers", 0.37, 0.38),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-FR", "France", "World Cup winner", 0.04, 0.05),
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-MEX", "Mexico", "Round of 16 Qualifiers", 0.65, 0.66),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-MEX", "Mexico", "Quarterfinals Qualifiers", 0.34, 0.35),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-MEX", "Mexico", "Semifinals Qualifiers", 0.16, None),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-MX", "Mexico", "World Cup winner", 0.01, 0.02),
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-AUS", "Australia", "Round of 16 Qualifiers", 0.44, 0.45),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-AUS", "Australia", "Quarterfinals Qualifiers", 0.24, 0.25),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-AUS", "Australia", "Semifinals Qualifiers", 0.09, 0.10),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-AU", "Australia", "World Cup winner", 0.07, 0.08),
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-ENG", "England", "Round of 16 Qualifiers", 0.71, 0.72),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-ENG", "England", "Quarterfinals Qualifiers", 0.49, 0.50),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-ENG", "England", "Semifinals Qualifiers", 0.35, 0.36),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-EN", "England", "World Cup winner", 0.29, 0.30),
        futures_market_row("KXWCROUND-26RO16", "KXWCROUND-26RO16-GER", "Germany", "Round of 16 Qualifiers", 0.69, 0.70),
        futures_market_row("KXWCROUND-26QUAR", "KXWCROUND-26QUAR-GER", "Germany", "Quarterfinals Qualifiers", 0.47, 0.48),
        futures_market_row("KXWCROUND-26SEMI", "KXWCROUND-26SEMI-GER", "Germany", "Semifinals Qualifiers", 0.33, 0.34),
        futures_market_row("KXMENWORLDCUP-26", "KXMENWORLDCUP-26-DE", "Germany", "World Cup winner", 0.00, 0.00),
    ]
    rows = pd.concat([rows, pd.DataFrame(proxy_rows)], ignore_index=True)
    rows.to_excel(path, index=False)


def futures_market_row(event_ticker, market_ticker, outcome, game, yes_bid, yes_ask):
    return {
        "event_ticker": event_ticker,
        "game": game,
        "sub_title": "",
        "category": "Sports",
        "market_ticker": market_ticker,
        "outcome": outcome,
        "market_title": f"Will {outcome} advance or win?",
        "yes_bid_dollars": yes_bid,
        "yes_ask_dollars": yes_ask,
        "last_price_dollars": yes_ask,
        "volume_fp": 500,
        "open_interest_fp": 200,
        "close_time": "2026-07-19T23:59:00Z",
    }


def write_sample_futures_value_board(path):
    bet_sheet = pd.DataFrame(
        [
            {
                "Stage": "Champion",
                "Team": "ARG",
                "champion_proxy_candidate": "YES",
                "strong_champion_proxy": "YES",
                "Action": "BUY YES",
                "Silver": 0.241,
                "Market Price": 0.15,
                "Edge": 0.091,
                "ROI": 0.61,
                "Quarter Kelly": 0.03,
                "proxy_kelly_dollars": 50.33,
                "proxy_contracts": 645,
                "ladder_1_price": 0.23,
                "ladder_1_contracts": 193,
                "ladder_2_price": 0.34,
                "ladder_2_contracts": 193,
                "ladder_3_price": 0.49,
                "ladder_3_contracts": 129,
                "ladder_4_price": 0.68,
                "ladder_4_contracts": 130,
                "ladder_expected_return": 261.62,
                "ladder_expected_profit": 164.17,
                "Stake on $500": 13.38,
                "Bucket": "B",
                "Volume": 1000,
                "event_ticker": "KXMENWORLDCUP-26",
                "market_ticker": "KXMENWORLDCUP-26-AR",
                "Position": "Yes",
                "Current Action": "BUY YES",
                "Entry Price": 0.15,
                "Current Value Change": 0.00,
                "Stake": 1.80,
                "Status": "Open",
            },
            {
                "Stage": "Round of 16",
                "Team": "USA",
                "champion_proxy_candidate": "YES",
                "strong_champion_proxy": None,
                "Action": "BUY YES",
                "Silver": 0.72,
                "Market Price": 0.61,
                "Edge": 0.11,
                "ROI": 0.18,
                "Quarter Kelly": 0.07,
                "proxy_kelly_dollars": None,
                "proxy_contracts": None,
                "ladder_1_price": None,
                "ladder_1_contracts": None,
                "ladder_2_price": None,
                "ladder_2_contracts": None,
                "ladder_3_price": None,
                "ladder_3_contracts": None,
                "ladder_4_price": None,
                "ladder_4_contracts": None,
                "ladder_expected_return": None,
                "ladder_expected_profit": None,
                "Stake on $500": 35.26,
                "Bucket": "A",
                "Volume": 800,
                "event_ticker": "KXWCROUND-26RO16",
                "market_ticker": "KXWCROUND-26RO16-USA",
                "Position": "No",
                "Current Action": None,
                "Entry Price": None,
                "Current Value Change": None,
                "Stake": None,
                "Status": None,
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        bet_sheet.to_excel(writer, sheet_name="Bet Sheet", index=False)


def write_sample_wager_log(path):
    wagers = pd.DataFrame(
        [
            {
                "Market Ticker": "KXWCGAME-26JUN13-USAMEX-USA",
                "Action": "BUY YES",
                "Entry Price": 0.45,
                "Contracts": 20,
                "Status": "Open",
            }
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wagers.to_excel(writer, sheet_name="Wagers", index=False)


def write_sample_ladder_value_board(path):
    value_board = pd.DataFrame(
        [
            {
                "game": "USA vs Mexico",
                "market_title": "United States wins",
                "market_ticker": "KXWCGAME-26JUN13-USAMEX-USA",
                "yes_bid": 0.50,
                "yes_ask": 0.50,
                "silver_prob": 0.60,
            },
            {
                "game": "USA vs Mexico",
                "market_title": "Mexico wins",
                "market_ticker": "KXWCGAME-26JUN13-USAMEX-MEX",
                "yes_bid": 0.61,
                "yes_ask": 0.62,
                "silver_prob": 0.25,
            },
            {
                "game": "USA vs Mexico",
                "market_title": "Tie",
                "market_ticker": "KXWCGAME-26JUN13-USAMEX-TIE",
                "yes_bid": 0.12,
                "yes_ask": 0.13,
                "silver_prob": 0.15,
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        value_board.to_excel(writer, sheet_name="Value Board", index=False)


def write_duplicate_header_ladder_value_board(path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Value Board"
    worksheet.append([
        "market_ticker",
        "market_ticker",
        "game",
        "event_title",
        "market_title",
        "yes_bid",
        "yes_ask",
        "silver_prob",
        "silver_prob",
        "edge",
        "edge",
    ])
    worksheet.append([
        "KXWCGAME-26JUN13-USAMEX-USA",
        "duplicate ignored",
        "USA vs Mexico",
        "Duplicate Event",
        "United States wins",
        0.50,
        0.51,
        0.60,
        0.61,
        0.09,
        0.10,
    ])
    worksheet.append([
        "KXWCGAME-26JUN13-USAMEX-MEX",
        "duplicate ignored",
        "USA vs Mexico",
        "Duplicate Event",
        "Mexico wins",
        0.61,
        0.62,
        0.25,
        0.26,
        0.11,
        0.12,
    ])
    workbook.save(path)


def write_sample_ladder_wager_log(path):
    wagers = pd.DataFrame(
        [
            {
                "Market Ticker": "KXWCGAME-26JUN13-USAMEX-USA",
                "Action": "BUY YES",
                "Entry Price": 0.45,
                "Contracts": 20,
                "Exit Contracts": 0,
                "Remaining Contracts": 20,
                "Status": "Open",
            },
            {
                "Market Ticker": "KXWCGAME-26JUN13-USAMEX-MEX",
                "Action": "SELL YES",
                "Entry Price": 0.60,
                "Contracts": 10,
                "Exit Contracts": 0,
                "Remaining Contracts": 10,
                "Status": "Partially Closed",
            },
            {
                "Market Ticker": "KXWCGAME-26JUN13-USAMEX-TIE",
                "Action": "BUY YES",
                "Entry Price": 0.10,
                "Contracts": 3,
                "Exit Contracts": 3,
                "Remaining Contracts": 0,
                "Status": "Settled",
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wagers.to_excel(writer, sheet_name="Wagers", index=False)


def write_zero_remaining_ladder_wager_log(path):
    wagers = pd.DataFrame(
        [
            {
                "Market Ticker": "KXWCGAME-26JUN13-USAMEX-USA",
                "Action": "BUY YES",
                "Entry Price": 0.45,
                "Contracts": 20,
                "Exit Contracts": 15,
                "Remaining Contracts": 5,
                "Status": "Partially Closed",
            },
            {
                "Market Ticker": "KXWCGAME-26JUN13-USAMEX-MEX",
                "Action": "SELL YES",
                "Entry Price": 0.60,
                "Contracts": 10,
                "Exit Contracts": 10,
                "Remaining Contracts": 0,
                "Status": "Partially Closed",
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wagers.to_excel(writer, sheet_name="Wagers", index=False)


def write_sample_futures_wager_log(path):
    wagers = pd.DataFrame(
        [
            {
                "Market Ticker": "KXMENWORLDCUP-26-AR",
                "Action": "BUY YES",
                "Entry Price": 0.15,
                "Contracts": 12,
                "Status": "Open",
            },
            {
                "Market Ticker": "KXWCROUND-26RO16-USA",
                "Action": "SELL YES",
                "Entry Price": 0.65,
                "Contracts": 10,
                "Status": "Partially Closed",
            },
            {
                "Market Ticker": "KXWCROUND-26QUAR-USA",
                "Action": "BUY YES",
                "Entry Price": 0.43,
                "Contracts": 8,
                "Status": "Settled",
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wagers.to_excel(writer, sheet_name="Wagers", index=False)


def write_sample_value_board_lookup(path):
    bet_sheet = pd.DataFrame(
        [
            {
                "Date": "2026-06-13",
                "Match": "USA vs Mexico",
                "Outcome": "United States",
                "Action": "BUY YES",
                "Silver": 0.60,
                "Edge": 0.10,
                "Bucket": "A",
                "market_ticker": "KXWCGAME-26JUN13-USAMEX-USA",
            },
            {
                "Date": "2026-06-13",
                "Match": "USA vs Mexico",
                "Outcome": "Mexico",
                "Action": "SELL YES",
                "Silver": 0.15,
                "Edge": 0.10,
                "Bucket": "A",
                "market_ticker": "KXWCGAME-26JUN13-USAMEX-MEX",
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        bet_sheet.to_excel(writer, sheet_name="Bet Sheet", index=False)


def trade(date, direction, market_ticker, amount, price_cents, fee):
    return {
        "type": "Trade",
        "Original_Date": date,
        "Direction": direction,
        "Market_Ticker": market_ticker,
        "Amount_In_Dollars": amount,
        "Price_In_Cents": price_cents,
        "Fee_In_Dollars": fee,
        "Result": "",
    }


def settlement(date, market_ticker, result):
    return {
        "type": "Settlement",
        "Original_Date": date,
        "Direction": "",
        "Market_Ticker": market_ticker,
        "Amount_In_Dollars": 0,
        "Price_In_Cents": 0,
        "Fee_In_Dollars": 0,
        "Result": result,
    }


def row_for(wagers, market_ticker):
    return wagers[wagers["Market Ticker"] == market_ticker].iloc[0]


def read_portfolio_sections(path):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook["Portfolio"]
    sections = {}
    current_title = None
    current_rows = []

    for values in worksheet.iter_rows(values_only=True):
        row = list(values)
        if all(value is None for value in row):
            if current_title is not None:
                sections[current_title] = current_rows
                current_title = None
                current_rows = []
            continue

        first_cell = row[0]
        if current_title is None:
            current_title = first_cell
            current_rows = []
        else:
            current_rows.append(row)

    if current_title is not None:
        sections[current_title] = current_rows

    return sections


def frame_for_section(rows):
    header = rows[0]
    body = rows[1:]
    frame = pd.DataFrame(body, columns=header).dropna(how="all")
    index_column = header[0]
    return frame.set_index(index_column)


def wager_log_archives(project_dir):
    return list((project_dir / "archive" / "wager_logs").glob("*_World_Cup_Bet_Log.xlsx"))


def write_expanded_value_board_lookup(path):
    bet_sheet = pd.DataFrame(
        [
            lookup_row("KXWCGAME-26JUL01-AAAEEE-AAA", "Jul 1, 2026", "AAA vs EEE", "AAA", "BUY YES", 0.58, 0.08, "B"),
            lookup_row("KXWCGAME-26JUL02-BBBFFF-BBB", "07/02/2026", "BBB vs FFF", "BBB", "BUY YES", 0.48, 0.06, "C"),
            lookup_row("KXWCGAME-26JUL03-CCCGGG-CCC", "2026-07-03", "CCC vs GGG", "CCC", "SELL YES", 0.38, 0.12, "A"),
            lookup_row("KXWCGAME-26JUL04-DDDHHH-DDD", "2026-07-04", "DDD vs HHH", "DDD", "BUY YES", 0.55, 0.07, "B"),
            lookup_row("KXWCGAME-26AUG05-IIIJJJ-III", "2026-08-05", "Current Match", "Current Outcome", "BUY YES", 0.99, 0.01, "C"),
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        bet_sheet.to_excel(writer, sheet_name="Bet Sheet", index=False)


def write_date_format_value_board_lookup(path):
    bet_sheet = pd.DataFrame(
        [
            lookup_row("KXWCGAME-26JUN14-STRONE-STR", "6/14/26", "STR vs ONE", "STR", "BUY YES", 0.58, 0.08, "B"),
            lookup_row("KXWCGAME-26JUN14-TIMTWO-TIM", "2026-06-01", "TIM vs TWO", "TIM", "BUY YES", 0.48, 0.06, "C"),
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        bet_sheet.to_excel(writer, sheet_name="Bet Sheet", index=False)


def lookup_row(market_ticker, date, match, outcome, action, silver, edge, bucket):
    return {
        "Date": date,
        "Match": match,
        "Outcome": outcome,
        "Action": action,
        "Silver": silver,
        "Edge": edge,
        "Bucket": bucket,
        "market_ticker": market_ticker,
    }


def write_prior_wager_log_for_preservation(path):
    wagers = pd.DataFrame(
        [
            {
                "Date Placed": "2026-05-01 08:00:00",
                "Market Ticker": "KXWCGAME-26JUL01-AAAEEE-AAA",
                "Market Result": "yes",
                "Contract Won": "Yes",
            },
            {
                "Date Placed": "2026-05-01 09:00:00",
                "Match Date": "August 5, 2026",
                "Match": "Prior Match",
                "Outcome": "Prior Outcome",
                "Action": "BUY YES",
                "Market Ticker": "KXWCGAME-26AUG05-IIIJJJ-III",
                "Silver Probability": 0.42,
                "Edge": 0.08,
                "Bucket": "B",
                "Market Result": "",
                "Contract Won": "",
                "Contracts": 5,
                "Entry Price": 0.25,
                "Exit Price": "",
                "Fee In": 0.05,
                "Fee Out": "",
                "Status": "Open",
                "Closing Price": 0.64,
                "CLV": 0.14,
                "CLV %": 0.28,
                "Realized P/L": "",
            },
            {
                "Date Placed": "2026-05-02 09:00:00",
                "Market Ticker": "KXWCGAME-26JUL02-BBBFFF-BBB",
                "Market Result": "yes",
                "Contract Won": "Yes",
                "Closing Price": 0.45,
                "CLV": 0.01,
                "CLV %": 0.02,
            },
            {
                "Date Placed": "2026-05-03 09:00:00",
                "Market Ticker": "KXWCGAME-26JUL03-CCCGGG-CCC",
                "Closing Price": 0.50,
                "CLV": 0.01,
                "CLV %": 0.02,
            },
            {
                "Date Placed": "2026-05-04 09:00:00",
                "Market Ticker": "KXWCGAME-26JUL04-DDDHHH-DDD",
                "Closing Price": "",
                "CLV": 0.03,
                "CLV %": 0.0632,
            }
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wagers.to_excel(writer, sheet_name="Wagers", index=False)


def write_prior_timestamp_match_date_log(path):
    wagers = pd.DataFrame(
        [
            {
                "Date Placed": "2026-05-01 09:00:00",
                "Match Date": pd.Timestamp("2026-06-14"),
                "Market Ticker": "KXWCGAME-26JUN14-TIMTWO-TIM",
                "Match": "Prior Timestamp Match",
                "Outcome": "Prior Timestamp Outcome",
                "Action": "BUY YES",
            },
            {
                "Date Placed": "2026-05-02 09:00:00",
                "Match Date": "0001-06-12",
                "Market Ticker": "KXWCGAME-26JUN12-BADOLD-BAD",
                "Match": "Prior Invalid Date Match",
                "Outcome": "Prior Invalid Date Outcome",
                "Action": "BUY YES",
            }
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wagers.to_excel(writer, sheet_name="Wagers", index=False)


def write_sample_web_wager_log(path):
    wagers = pd.DataFrame(
        [
            {
                "Date Placed": "2026-06-10 09:00:00",
                "Market Type": "Match",
                "Stage": None,
                "Team": "USA",
                "Match Date": "2026-06-13",
                "Match": "USA vs Mexico",
                "Outcome": "United States",
                "Action": "BUY YES",
                "Silver Probability": 0.6000000000001,
                "Edge": 0.123456,
                "Bucket": "A",
                "Status": "Open",
                "Contract Won": None,
                "Entry Price": 0.456789,
                "Exit Price": None,
                "Closing Price": None,
                "CLV": None,
                "CLV %": None,
                "Realized P/L": None,
            },
            {
                "Date Placed": "2026-06-11 10:00:00",
                "Market Type": "Match",
                "Stage": None,
                "Team": "BRA",
                "Match Date": "6/14/26",
                "Match": "Canada vs Brazil",
                "Outcome": "Brazil",
                "Action": "SELL YES",
                "Silver Probability": 0.35,
                "Edge": 0.08,
                "Bucket": "B",
                "Status": "Settled",
                "Contract Won": "Yes",
                "Entry Price": 0.58,
                "Exit Price": "",
                "Closing Price": 0.501234,
                "CLV": 0.083333,
                "CLV %": 0.141592,
                "Realized P/L": 2.2,
            },
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wagers.to_excel(writer, sheet_name="Wagers", index=False)


if __name__ == "__main__":
    unittest.main()
